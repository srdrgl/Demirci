
#main.py
# civileng.serdar@gmail.com 
"""
Demirci - Smart Rebar Cutting Optimization
Integrated with Lexicographic Optimization Algorithm

"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter import font as tkfont
from datetime import datetime
from typing import List, Dict
import os
import webbrowser

# GitHub Profile
GITHUB_PROFILE = "https://github.com/srdrgl"

# Import optimization functions
from calculations import solve_multi_diameter_lexicographic


class RebarOptimizerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Demirci - Smart Rebar Cutting Optimization v1.0")
        
        # Set window icon (cross-platform - using PNG for all platforms)
        try:
            self.icon = tk.PhotoImage(file='demirci_icon.png')
            self.root.iconphoto(True, self.icon)
        except Exception as e:
            pass  # Icon bulunamazsa program yine de çalışsın
        
        # Set initial window size to fit content
        window_width = 1250
        window_height = 800
        
        # Center window on screen
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Minimum size
        self.root.minsize(1100, 720)
        
        # Professional color palette
        self.colors = {
            'primary': '#061E29',
            'primary_light': '#1D546D',
            'secondary': '#5F9598',
            'accent': '#1D546D',
            'accent_light': '#5F9598',
            'success': '#27AE60',
            'success_dark': '#229954',
            'warning': '#F39C12',
            'danger': '#C0392B',
            'light': '#F3F4F4',
            'light_gray': '#E8EDEF',
            'dark': '#061E29',
            'dark_gray': '#7F8C8D',
            'white': '#FFFFFF',
            'text_dark': '#061E29',
            'text_light': '#95A5A6',
            'card_bg': '#FFFFFF',
            'border': '#BDC3C7',
            'steel_blue': '#1D546D',
            'warm_bg': '#F8F9F9'
        }
        
        # Modern font settings
        self.title_font = tkfont.Font(family="Helvetica", size=22, weight="bold")
        self.header_font = tkfont.Font(family="Helvetica", size=12, weight="bold")
        self.normal_font = tkfont.Font(family="Helvetica", size=9)
        self.small_font = tkfont.Font(family="Helvetica", size=8)
        
        # Data storage
        self.rebar_list = []
        self.stock_length = 12.0  # meters
        self.optimization_results = None
        
        self.setup_ui()
        
    def setup_ui(self):
        """Create main UI components with modern styling"""
        # Set window background
        self.root.configure(bg=self.colors['light_gray'])
        
        # Modern style configuration
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure modern styles
        style.configure('TFrame', background=self.colors['light_gray'])
        style.configure('TLabel', background=self.colors['white'], 
                       foreground=self.colors['text_dark'], font=self.normal_font)
        style.configure('TLabelframe', background=self.colors['white'], 
                       foreground=self.colors['text_dark'], borderwidth=0, relief='flat')
        style.configure('TLabelframe.Label', font=self.header_font, 
                       foreground=self.colors['secondary'])
        
        # Header panel
        self.create_header()
        
        # Main content container
        content_wrapper = tk.Frame(self.root, bg=self.colors['light_gray'])
        content_wrapper.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Main content area
        main_container = ttk.PanedWindow(content_wrapper, orient=tk.HORIZONTAL)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Left panel - Data input
        left_panel = self.create_input_panel(main_container)
        main_container.add(left_panel, weight=2)
        
        # Right panel - Results
        right_panel = self.create_results_panel(main_container)
        main_container.add(right_panel, weight=3)
        
        # Bottom status bar
        self.create_status_bar()
    
    def create_header(self):
        """Create compact modern header with professional colors"""
        # Main header frame - height artırıldı
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=70)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Main container for title and subtitle
        main_container = tk.Frame(header_frame, bg=self.colors['primary'])
        main_container.pack(side=tk.LEFT, padx=20, pady=15)
        
        # Icon + Title on the left
        try:
            # Load icon image
            icon_img = tk.PhotoImage(file='demirci_icon.png')
            # Resize to 54x54 px (512/9.5 ≈ 54, original 32'nin %70 büyüğü)
            icon_img = icon_img.subsample(3, 3)  # %70 daha büyük
            
            icon_label = tk.Label(
                main_container,
                image=icon_img,
                bg=self.colors['primary']
            )
            icon_label.image = icon_img  # Keep reference
            icon_label.pack(side=tk.LEFT, padx=(0, 10))
        except:
            # Fallback to emoji if image not found
            pass
        
        # Title
        title_label = tk.Label(
            main_container,
            text="Demirci",  # Emoji kaldırıldı
            font=self.title_font,
            bg=self.colors['primary'],
            fg=self.colors['white']
        )
        title_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Subtitle on the right
        subtitle_label = tk.Label(
            main_container,
            text="Rebar Cutting Optimizer",
            font=self.normal_font,
            bg=self.colors['primary'],
            fg=self.colors['light']
        )
        subtitle_label.pack(side=tk.LEFT)
        
        # Right side - Contact button
        right_container = tk.Frame(header_frame, bg=self.colors['primary'])
        right_container.pack(side=tk.RIGHT, padx=20, pady=15)
        
        contact_btn = tk.Button(
            right_container,
            text="📧 Contact",
            command=self.open_contact,
            bg=self.colors['accent_light'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=8,
            cursor="hand2",
            activebackground=self.colors['secondary']
        )
        contact_btn.pack()
        
        # Decorative line
        line_frame = tk.Frame(self.root, bg=self.colors['secondary'], height=3)
        line_frame.pack(fill=tk.X)
    
    def open_contact(self):
        """Open GitHub Profile for contact"""
        try:
            webbrowser.open(GITHUB_PROFILE)
            self.status_label.config(text="● Opening GitHub Profile...")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open browser:\n{str(e)}")
    
    def create_input_panel(self, parent):
        """Left panel - Compact data input form"""
        # Main panel with card effect
        panel_wrapper = tk.Frame(parent, bg=self.colors['light_gray'])
        panel = tk.Frame(panel_wrapper, bg=self.colors['card_bg'], relief='flat', bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=(0, 5), pady=0)
        
        # Add padding inside the card
        inner_panel = tk.Frame(panel, bg=self.colors['card_bg'])
        inner_panel.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # Modern title with icon
        title_frame = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = tk.Label(
            title_frame,
            text="📋 Rebar Information",
            font=self.header_font,
            bg=self.colors['card_bg'],
            fg=self.colors['secondary'],
            anchor=tk.W
        )
        title_label.pack(side=tk.LEFT)
        
        # Settings section
        settings_frame = tk.LabelFrame(
            inner_panel, 
            text="⚙️ Settings", 
            font=self.header_font,
            bg=self.colors['warm_bg'],
            fg=self.colors['secondary'],
            relief='flat',
            bd=1,
            padx=10,
            pady=8
        )
        settings_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(
            settings_frame, 
            text="Stock Bar Length (m):",
            bg=self.colors['warm_bg'],
            fg=self.colors['text_dark'],
            font=self.normal_font
        ).grid(row=0, column=0, sticky=tk.W, pady=3)
        
        self.stock_entry = tk.Entry(
            settings_frame, 
            width=10,
            font=self.normal_font,
            relief='flat',
            bd=2,
            bg=self.colors['white']
        )
        self.stock_entry.insert(0, "12.0")
        self.stock_entry.grid(row=0, column=1, padx=8, pady=3)
        
        # Add rebar form
        input_frame = tk.LabelFrame(
            inner_panel, 
            text="➕ Add New Rebar", 
            font=self.header_font,
            bg=self.colors['warm_bg'],
            fg=self.colors['secondary'],
            relief='flat',
            bd=1,
            padx=10,
            pady=8
        )
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Form fields
        fields = [
            ("Diameter (mm):", 0),
            ("Length (m):", 1),
            ("Quantity:", 2)
        ]
        
        for label_text, row in fields:
            tk.Label(
                input_frame, 
                text=label_text,
                bg=self.colors['warm_bg'],
                fg=self.colors['text_dark'],
                font=self.normal_font
            ).grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0, 8))
        
        # Diameter combo
        self.diameter_var = tk.StringVar()
        self.diameter_combo = ttk.Combobox(
            input_frame,
            textvariable=self.diameter_var,
            values=["8", "10", "12", "14", "16", "18", "20", "22", "24", "28", "32", "40", "50"],
            width=16,
            state="readonly",
            font=self.normal_font
        )
        self.diameter_combo.current(0)
        self.diameter_combo.grid(row=0, column=1, pady=5, sticky=tk.W)
        
        # Length entry
        self.length_entry = tk.Entry(
            input_frame, 
            width=18,
            font=self.normal_font,
            relief='flat',
            bd=2,
            bg=self.colors['white']
        )
        self.length_entry.grid(row=1, column=1, pady=5, sticky=tk.W)
        
        # Quantity entry
        self.quantity_entry = tk.Entry(
            input_frame, 
            width=18,
            font=self.normal_font,
            relief='flat',
            bd=2,
            bg=self.colors['white']
        )
        self.quantity_entry.grid(row=2, column=1, pady=5, sticky=tk.W)
        
        # Compact buttons
        btn_frame = tk.Frame(input_frame, bg=self.colors['warm_bg'])
        btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
        
        add_btn = tk.Button(
            btn_frame,
            text="➕ Add",
            command=self.add_rebar,
            bg=self.colors['success'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=6,
            cursor="hand2"
        )
        add_btn.pack(side=tk.LEFT, padx=3)
        
        excel_btn = tk.Button(
            btn_frame,
            text="📁 Excel",
            command=self.load_excel,
            bg=self.colors['steel_blue'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=6,
            cursor="hand2"
        )
        excel_btn.pack(side=tk.LEFT, padx=3)
        
        # Rebar list
        list_frame = tk.LabelFrame(
            inner_panel, 
            text="📊 Added Rebars", 
            font=self.header_font,
            bg=self.colors['warm_bg'],
            fg=self.colors['secondary'],
            relief='flat',
            bd=1,
            padx=8,
            pady=8
        )
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Treeview with compact style
        columns = ("Diameter", "Length", "Quantity")
        self.tree = ttk.Treeview(
            list_frame, 
            columns=columns, 
            show="headings", 
            height=10,
            selectmode='extended'
        )
        
        # Configure column headings
        col_widths = {"Diameter": 70, "Length": 80, "Quantity": 70}
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths[col])
        
        # Configure tree colors
        style = ttk.Style()
        style.configure("Treeview",
                       background=self.colors['white'],
                       foreground=self.colors['text_dark'],
                       fieldbackground=self.colors['white'],
                       font=self.normal_font,
                       rowheight=22)
        style.configure("Treeview.Heading",
                       background=self.colors['secondary'],
                       foreground=self.colors['white'],
                       font=self.normal_font)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # List action buttons
        list_btn_frame = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        list_btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        delete_btn = tk.Button(
            list_btn_frame,
            text="❌ Delete",
            command=self.delete_rebar,
            bg=self.colors['danger'],
            fg=self.colors['white'],
            font=self.small_font,
            relief='flat',
            bd=0,
            padx=12,
            pady=5,
            cursor="hand2"
        )
        delete_btn.pack(side=tk.LEFT, padx=3)
        
        clear_btn = tk.Button(
            list_btn_frame,
            text="🗑️ Clear All",
            command=self.clear_all,
            bg=self.colors['dark_gray'],
            fg=self.colors['white'],
            font=self.small_font,
            relief='flat',
            bd=0,
            padx=12,
            pady=5,
            cursor="hand2"
        )
        clear_btn.pack(side=tk.LEFT, padx=3)
        
        # Calculate button
        calculate_btn = tk.Button(
            inner_panel,
            text="⚡ CALCULATE",
            command=self.calculate_optimization,
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            font=self.header_font,
            relief='flat',
            bd=0,
            height=2,
            cursor="hand2",
            activebackground=self.colors['accent']
        )
        calculate_btn.pack(fill=tk.X, pady=0)
        
        return panel_wrapper
    
    def create_results_panel(self, parent):
        """Right panel - Results display"""
        # Main panel with card effect
        panel_wrapper = tk.Frame(parent, bg=self.colors['light_gray'])
        panel = tk.Frame(panel_wrapper, bg=self.colors['card_bg'], relief='flat', bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=(5, 0), pady=0)
        
        # Add padding inside the card
        inner_panel = tk.Frame(panel, bg=self.colors['card_bg'])
        inner_panel.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # Modern title
        title_frame = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = tk.Label(
            title_frame,
            text="📈 Optimization Results",
            font=self.header_font,
            bg=self.colors['card_bg'],
            fg=self.colors['secondary'],
            anchor=tk.W
        )
        title_label.pack(side=tk.LEFT)
        
        # Summary information
        summary_container = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        summary_container.pack(fill=tk.X, pady=(0, 10))
        
        self.summary_labels = {}
        summary_items = [
            ("Total Demand", "demand", "📦"),
            ("Bars Used", "bars", "📊"),
            ("Total Waste", "waste", "🗑️"),
            ("Waste %", "waste_pct", "📉")
        ]
        
        for i, (label, key, icon) in enumerate(summary_items):
            card = tk.Frame(summary_container, bg=self.colors['warm_bg'], relief='flat', bd=1)
            card.pack(fill=tk.X, pady=2)
            
            # Left side - icon and label
            left_frame = tk.Frame(card, bg=self.colors['warm_bg'])
            left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8, pady=6)
            
            label_text = tk.Label(
                left_frame,
                text=f"{icon} {label}:",
                font=self.normal_font,
                bg=self.colors['warm_bg'],
                fg=self.colors['text_dark'],
                anchor=tk.W
            )
            label_text.pack(side=tk.LEFT)
            
            # Right side - value
            self.summary_labels[key] = tk.Label(
                card,
                text="-",
                font=self.header_font,
                bg=self.colors['warm_bg'],
                fg=self.colors['secondary'],
                anchor=tk.E,
                padx=8
            )
            self.summary_labels[key].pack(side=tk.RIGHT, pady=6)
        
        # Cutting plan
        plan_label_frame = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        plan_label_frame.pack(fill=tk.X, pady=(8, 5))
        
        tk.Label(
            plan_label_frame,
            text="✂️ Cutting Plan",
            font=self.header_font,
            bg=self.colors['card_bg'],
            fg=self.colors['secondary'],
            anchor=tk.W
        ).pack(side=tk.LEFT)
        
        plan_frame = tk.Frame(inner_panel, bg=self.colors['warm_bg'], relief='flat', bd=1)
        plan_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Text widget
        self.results_text = tk.Text(
            plan_frame,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg=self.colors['white'],
            fg=self.colors['text_dark'],
            relief='flat',
            bd=0,
            padx=12,
            pady=12
        )
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar_frame = tk.Frame(plan_frame, bg=self.colors['warm_bg'], width=12)
        scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        results_scroll = tk.Scrollbar(
            scrollbar_frame, 
            orient=tk.VERTICAL, 
            command=self.results_text.yview,
            relief='flat',
            bg=self.colors['warm_bg']
        )
        results_scroll.pack(fill=tk.Y, expand=True)
        self.results_text.configure(yscrollcommand=results_scroll.set)
        
        # Action buttons
        action_frame = tk.Frame(inner_panel, bg=self.colors['card_bg'])
        action_frame.pack(fill=tk.X)
        
        save_btn = tk.Button(
            action_frame,
            text="💾 Save",
            command=self.save_report,
            bg=self.colors['success'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=6,
            cursor="hand2"
        )
        save_btn.pack(side=tk.LEFT, padx=3)
        
        print_btn = tk.Button(
            action_frame,
            text="🖨️ Print",
            command=self.print_report,
            bg=self.colors['steel_blue'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=6,
            cursor="hand2"
        )
        print_btn.pack(side=tk.LEFT, padx=3)
        
        copy_btn = tk.Button(
            action_frame,
            text="📋 Copy",
            command=self.copy_to_clipboard,
            bg=self.colors['dark_gray'],
            fg=self.colors['white'],
            font=self.normal_font,
            relief='flat',
            bd=0,
            padx=15,
            pady=6,
            cursor="hand2"
        )
        copy_btn.pack(side=tk.LEFT, padx=3)
        
        return panel_wrapper
    
    def create_status_bar(self):
        """Bottom status bar"""
        status_frame = tk.Frame(self.root, bg=self.colors['primary'], height=28)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        status_frame.pack_propagate(False)
        
        # Status icon and text
        status_container = tk.Frame(status_frame, bg=self.colors['primary'])
        status_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12)
        
        self.status_label = tk.Label(
            status_container,
            text="● Ready",
            bg=self.colors['primary'],
            fg=self.colors['white'],
            anchor=tk.W,
            font=self.normal_font
        )
        self.status_label.pack(side=tk.LEFT, pady=6)
        
        # Version and info
        info_frame = tk.Frame(status_frame, bg=self.colors['primary'])
        info_frame.pack(side=tk.RIGHT, padx=12)
        
        version_label = tk.Label(
            info_frame,
            text="v1.0  |  Demirci",
            bg=self.colors['primary'],
            fg=self.colors['text_light'],
            font=self.small_font
        )
        version_label.pack(pady=6)
    
    def add_rebar(self):
        """Add new rebar to list"""
        try:
            diameter = self.diameter_var.get()
            length = float(self.length_entry.get().replace(',', '.'))
            quantity = int(self.quantity_entry.get())
            
            if length <= 0 or quantity <= 0:
                raise ValueError("Length and quantity must be positive!")
            
            if length > self.stock_length:
                raise ValueError(f"Rebar length cannot exceed stock length ({self.stock_length}m)!")
            
            # Add to list
            rebar_info = {
                'diameter': int(diameter),
                'length': length,
                'quantity': quantity
            }
            self.rebar_list.append(rebar_info)
            
            # Add to TreeView
            self.tree.insert(
                "",
                tk.END,
                values=(
                    f"Ø{diameter}",
                    f"{length:.2f}m",
                    quantity
                )
            )
            
            # Clear form
            self.length_entry.delete(0, tk.END)
            self.quantity_entry.delete(0, tk.END)
            
            self.status_label.config(text=f"● Rebar added: Ø{diameter} - {length}m x {quantity} pcs")
            
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid value: {str(e)}")
    
    def delete_rebar(self):
        """Delete selected rebar"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select rebar to delete!")
            return
        
        # Sort indices in reverse to delete from end
        indices = sorted([self.tree.index(item) for item in selected], reverse=True)
        
        for index in indices:
            item = self.tree.get_children()[index]
            self.tree.delete(item)
            del self.rebar_list[index]
        
        self.status_label.config(text="● Selected rebar(s) deleted")
    
    def clear_all(self):
        """Clear all list"""
        if messagebox.askyesno("Confirm", "Are you sure you want to clear all rebars?"):
            self.tree.delete(*self.tree.get_children())
            self.rebar_list.clear()
            self.status_label.config(text="● List cleared")
    
    def load_excel(self):
        """Load rebar list from Excel file using smart reader"""
        filename = filedialog.askopenfilename(
            title="Select File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("ODS files", "*.ods"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            try:
                # Use smart reader function
                demands = self.read_file_to_demands(filename)
                
                # Clear existing list
                self.tree.delete(*self.tree.get_children())
                self.rebar_list.clear()
                
                # Add to GUI
                total_added = 0
                for diameter in sorted(demands.keys()):
                    for length, count in zip(demands[diameter]['lengths'], demands[diameter]['counts']):
                        # Add to internal list
                        rebar_info = {
                            'diameter': diameter,
                            'length': length,
                            'quantity': count
                        }
                        self.rebar_list.append(rebar_info)
                        
                        # Add to TreeView
                        self.tree.insert(
                            "",
                            tk.END,
                            values=(
                                f"Ø{diameter}",
                                f"{length:.2f}m",
                                count
                            )
                        )
                        total_added += 1
                
                self.status_label.config(text=f"● File loaded: {total_added} items added")
                messagebox.showinfo("Success", f"File loaded successfully!\n{total_added} items added")
                
            except ImportError as e:
                messagebox.showerror(
                    "Error",
                    "Required libraries not found!\n\n"
                    "Install with: pip install pandas openpyxl odfpy"
                )
            except Exception as e:
                messagebox.showerror("Error", f"Could not read file:\n{str(e)}")
    
    def read_file_to_demands(self, file_path: str) -> Dict[int, Dict[str, List]]:
        """
        Smart file reader - Auto-detects columns and formats
        Supports: .xlsx, .xls, .ods, .csv
        UPDATED: Now uses partial matching for column names (e.g., "çap (mm)" matches "çap")
        """
        import pandas as pd
        from collections import defaultdict
        
        # Column name aliases
        diameter_aliases = [
            'çap', 'cap', 'çaplar', 'caplar',
            'diameter', 'diameters', 'dia',
            'kalınlık', 'kalinlik'
        ]
        
        length_aliases = [
            'uzunluk', 'uzunluklar', 'boy', 'boylar',
            'length', 'lengths', 'len',
            'mesafe', 'metre', 'meter'
        ]
        
        count_aliases = [
            'adet', 'adetler', 'miktar', 'miktarlar',
            'sayı', 'sayılar', 'sayi', 'sayilar',
            'count', 'counts', 'quantity', 'quantities',
            'qty', 'number', 'adet/miktar'
        ]
        
        # Read file based on format
        if file_path.endswith('.csv'):
            try:
                df = pd.read_csv(file_path, encoding='utf-8')
            except:
                df = pd.read_csv(file_path, encoding='latin-1')
        else:
            engine = None
            if file_path.endswith('.ods'):
                engine = 'odf'
            
            # Read all rows first (no header)
            df_raw = pd.read_excel(file_path, sheet_name=0, engine=engine, header=None)
            
            # Find header row (first non-empty row with at least 2 cells)
            header_row = None
            for idx, row in df_raw.iterrows():
                non_empty = row.dropna()
                if len(non_empty) >= 2:
                    header_row = idx
                    break
            
            if header_row is None:
                raise ValueError("Header row not found in file!")
            
            # Read again from header row
            df = pd.read_excel(file_path, sheet_name=0, engine=engine, header=header_row)
        
        # Normalize column names
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        # Find columns using PARTIAL MATCHING
        diameter_col = None
        length_col = None
        count_col = None
        
        for col in df.columns:
            col_clean = col.replace(' ', '').replace('_', '')
            
            # UPDATED: Use 'any' with 'in' for partial matching
            if not diameter_col and any(alias in col_clean for alias in diameter_aliases):
                diameter_col = col
            
            if not length_col and any(alias in col_clean for alias in length_aliases):
                length_col = col
            
            if not count_col and any(alias in col_clean for alias in count_aliases):
                count_col = col
        
        # Check required columns
        if not diameter_col:
            raise ValueError(
                f"Diameter column not found!\n"
                f"Available columns: {list(df.columns)}\n"
                f"Accepted names: {', '.join(diameter_aliases[:5])}..."
            )
        
        if not length_col:
            raise ValueError(
                f"Length column not found!\n"
                f"Available columns: {list(df.columns)}\n"
                f"Accepted names: {', '.join(length_aliases[:5])}..."
            )
        
        # If no count column, assume 1 per row
        if not count_col:
            df['_temp_count'] = 1
            count_col = '_temp_count'
        
        # Remove empty rows
        df = df.dropna(subset=[diameter_col, length_col])
        
        # Clean numbers (replace comma with dot)
        def clean_number(val):
            if pd.isna(val):
                return None
            val_str = str(val).replace(',', '.')
            try:
                return float(val_str)
            except:
                return None
        
        df[diameter_col] = df[diameter_col].apply(clean_number)
        df[length_col] = df[length_col].apply(clean_number)
        df[count_col] = df[count_col].apply(clean_number)
        
        # Remove NaN rows
        df = df.dropna(subset=[diameter_col, length_col, count_col])
        
        # Convert types
        df[diameter_col] = df[diameter_col].astype(int)
        df[count_col] = df[count_col].astype(int)
        
        # Build demands dictionary
        demands = defaultdict(lambda: {'lengths': [], 'counts': []})
        seen = {}  # (diameter, length) -> index
        
        for _, row in df.iterrows():
            diameter = int(row[diameter_col])
            length = float(row[length_col])
            count = int(row[count_col])
            
            key = (diameter, length)
            
            # If same diameter+length seen before, add counts
            if key in seen:
                idx = seen[key]
                demands[diameter]['counts'][idx] += count
            else:
                # First time, add new entry
                demands[diameter]['lengths'].append(length)
                demands[diameter]['counts'].append(count)
                seen[key] = len(demands[diameter]['lengths']) - 1
        
        return dict(demands)
    
    def calculate_optimization(self):
        """Calculate optimization using the fixed algorithm"""
        if not self.rebar_list:
            messagebox.showwarning("Warning", "Please add rebars first!")
            return
        
        self.status_label.config(text="⚡ Calculating optimization...")
        self.root.update()
        
        try:
            # Update stock length
            self.stock_length = float(self.stock_entry.get().replace(',', '.'))
            
            # Prepare data for multi-diameter optimization
            demands = {}
            for rebar in self.rebar_list:
                diameter = rebar['diameter']
                if diameter not in demands:
                    demands[diameter] = {'lengths': [], 'counts': []}
                
                demands[diameter]['lengths'].append(rebar['length'])
                demands[diameter]['counts'].append(rebar['quantity'])
            
            # Run optimization (without console output)
            self.optimization_results = solve_multi_diameter_lexicographic(
                demands=demands,
                bin_capacity=self.stock_length,
                min_efficiency=0.85,
                max_patterns=1000,
                phase1_time_limit_ms=90000,
                phase2_time_limit_ms=90000,
                verbose=False,  # No console output
                print_output=False,  # No console printing
                adaptive=True
            )
            
            # Display results in GUI
            self.display_optimization_results()
            
            self.status_label.config(text="● Calculation completed!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Calculation error:\n{str(e)}")
            self.status_label.config(text="✗ Calculation failed!")
            import traceback
            traceback.print_exc()
    
    def display_optimization_results(self):
        """Display optimization results in GUI"""
        if not self.optimization_results:
            return
        
        # Calculate totals
        total_bars = 0
        total_waste = 0
        total_demand = 0
        total_capacity = 0
        theoretical_total = 0
        
        for diameter, result in self.optimization_results.items():
            if result:
                total_bars += result['total_bins']
                total_waste += result['total_waste']
                total_demand += result['total_demand']
                total_capacity += result['total_capacity']
                theoretical_total += result['theoretical_min']
        
        # Update summary labels
        overall_waste_pct = (total_waste / total_capacity * 100) if total_capacity > 0 else 0
        
        self.summary_labels['demand'].config(text=f"{total_demand:.2f}m")
        self.summary_labels['bars'].config(text=f"{theoretical_total}/{total_bars} bars")  # Theoretical/Used
        self.summary_labels['waste'].config(text=f"{total_waste:.2f}m")
        self.summary_labels['waste_pct'].config(
            text=f"{overall_waste_pct:.2f}%",
            foreground=self.colors['success'] if overall_waste_pct < 10 else self.colors['warning']
        )
        
        # Display cutting plan
        self.results_text.delete(1.0, tk.END)
        
        self.results_text.insert(tk.END, "=" * 80 + "\n")
        self.results_text.insert(tk.END, "CUTTING PLAN - PRODUCTION INSTRUCTION\n")
        self.results_text.insert(tk.END, "=" * 80 + "\n\n")
        
        self.results_text.insert(tk.END, f"Date: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n")
        self.results_text.insert(tk.END, f"Stock Bar Length: {self.stock_length}m\n\n")
        
        # SUMMARY TABLE
        self.results_text.insert(tk.END, "=" * 85 + "\n")
        self.results_text.insert(tk.END, "OVERALL SUMMARY - ALL DIAMETERS\n")
        self.results_text.insert(tk.END, "=" * 85 + "\n\n")
        
        # Table header
        self.results_text.insert(tk.END, f"{'DIAM(mm)':<12} {'BARS':<10} {'WASTE(m)':<12} {'WASTE%':<10} {'DEMAND(m)':<12}\n")
        self.results_text.insert(tk.END, "-" * 60 + "\n")
        
        # Table rows
        for diameter in sorted(self.optimization_results.keys()):
            result = self.optimization_results[diameter]
            if result:
                self.results_text.insert(
                    tk.END,
                    f"{diameter:<12} {result['total_bins']:<10} "
                    f"{result['total_waste']:<12.2f} "
                    f"{result['waste_percentage']:<10.2f} "
                    f"{result['total_demand']:<12.2f}\n"
                )
            else:
                self.results_text.insert(tk.END, f"{diameter:<12} {'NO SOLUTION':<10}\n")
        
        self.results_text.insert(tk.END, "-" * 60 + "\n")
        self.results_text.insert(
            tk.END,
            f"{'TOTAL':<12} {total_bars:<10} "
            f"{total_waste:<12.2f} "
            f"{overall_waste_pct:<10.2f} "
            f"{total_demand:<12.2f}\n"
        )
        self.results_text.insert(tk.END, "=" * 85 + "\n\n")
        
        # Display patterns for each diameter
        for diameter in sorted(self.optimization_results.keys()):
            result = self.optimization_results[diameter]
            
            if not result:
                self.results_text.insert(tk.END, f"\n{'='*80}\n")
                self.results_text.insert(tk.END, f"DIAMETER: Ø{diameter}mm\n")
                self.results_text.insert(tk.END, f"{'='*80}\n")
                self.results_text.insert(tk.END, "⚠️ NO SOLUTION FOUND\n\n")
                continue
            
            self.results_text.insert(tk.END, f"\n{'='*80}\n")
            self.results_text.insert(tk.END, f"DIAMETER: Ø{diameter}mm\n")
            self.results_text.insert(tk.END, f"{'='*80}\n\n")
            
            self.results_text.insert(tk.END, f"Demand: {result['total_demand']:.2f}m\n")
            self.results_text.insert(tk.END, f"Bars (Theoretical/Used): {result['theoretical_min']}/{result['total_bins']} bars\n")
            self.results_text.insert(tk.END, f"Waste: {result['total_waste']:.2f}m ({result['waste_percentage']:.2f}%)\n\n")
            
            self.results_text.insert(tk.END, "-" * 80 + "\n")
            self.results_text.insert(tk.END, "CUTTING PATTERNS:\n")
            self.results_text.insert(tk.END, "-" * 80 + "\n\n")
            
            # Get lengths for this diameter
            diameter_rebars = [r for r in self.rebar_list if r['diameter'] == diameter]
            lengths = [r['length'] for r in diameter_rebars]
            
            # Display patterns
            for idx, pattern_data in enumerate(result['used_patterns'], 1):
                combo = pattern_data['combo']
                count = pattern_data['count']
                total = pattern_data['total']
                waste = pattern_data['waste']
                utilization = (total / self.stock_length) * 100
                
                # Build pattern description
                cuts = []
                for i, pieces in enumerate(combo):
                    if pieces > 0:
                        cuts.append(f"{pieces}×{lengths[i]:.2f}m")
                
                pattern_desc = " + ".join(cuts)
                
                self.results_text.insert(tk.END, f"Pattern {idx}: {count} bars\n")
                self.results_text.insert(tk.END, f"  Cuts: {pattern_desc}\n")
                self.results_text.insert(tk.END, f"  Total: {total:.2f}m | Waste: {waste:.2f}m | Utilization: {utilization:.1f}%\n\n")
        
        self.results_text.insert(tk.END, "=" * 80 + "\n")
        self.results_text.insert(tk.END, "⚠️ NOTE: Double-check all measurements before cutting.\n")
        self.results_text.insert(tk.END, "=" * 80 + "\n")
        self.results_text.insert(tk.END, "END OF CUTTING PLAN\n")
        self.results_text.insert(tk.END, "=" * 80 + "\n")
    
    def save_report(self):
        """Save report to file (TXT, Excel, or PDF)"""
        if not self.optimization_results:
            messagebox.showwarning("Warning", "No report to save!")
            return
        
        # File types
        file_types = [
            ("Text files", "*.txt"),
            ("Excel files", "*.xlsx"),
            ("PDF files", "*.pdf")
        ]
        
        # Default name
        default_name = f"cutting_plan_{datetime.now().strftime('%Y%m%d_%H%M')}"
        
        # File save dialog
        filename = filedialog.asksaveasfilename(
            title="Save Report As",
            initialfile=default_name,
            filetypes=file_types,
            defaultextension=""
        )
        
        if not filename:
            return
        
        # Check extension and call appropriate function
        if not (filename.endswith('.txt') or filename.endswith('.xlsx') or filename.endswith('.pdf')):
            if '.txt' in str(file_types[0]):
                filename += '.txt'
            elif '.xlsx' in str(file_types[1]):
                filename += '.xlsx'
            elif '.pdf' in str(file_types[2]):
                filename += '.pdf'
        
        # Call appropriate save function
        if filename.lower().endswith('.xlsx'):
            self.save_as_excel(filename)
        elif filename.lower().endswith('.pdf'):
            self.save_as_pdf(filename)
        else:
            if not filename.lower().endswith('.txt'):
                filename += '.txt'
            self.save_as_text(filename)
    
    def save_as_text(self, filename):
        """Save report as text file"""
        content = self.results_text.get(1.0, tk.END)
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(content)
            messagebox.showinfo("Success", f"Text file saved:\n{filename}")
            self.status_label.config(text=f"● Text saved: {os.path.basename(filename)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file:\n{str(e)}")
    
    def save_as_excel(self, filename):
        """Save report as Excel file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
        except ImportError:
            messagebox.showerror(
                "Error",
                "Excel export requires pandas and openpyxl!\n\n"
                "Install with: pip install pandas openpyxl"
            )
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cutting Plan"
            
            # Styles
            header_fill = PatternFill(start_color="5F9598", end_color="5F9598", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = "CUTTING PLAN - PRODUCTION INSTRUCTION"
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            row += 2
            
            # Date and settings
            ws[f'A{row}'] = f"Date: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            row += 1
            ws[f'A{row}'] = f"Stock Bar Length: {self.stock_length}m"
            row += 2
            
            # Summary table header
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = "OVERALL SUMMARY - ALL DIAMETERS"
            cell.font = Font(bold=True, size=12)
            row += 1
            
            # Column headers
            headers = ['DIAM(mm)', 'BARS', 'WASTE(m)', 'WASTE%', 'DEMAND(m)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            row += 1
            
            # Calculate totals
            total_bars = 0
            total_waste = 0
            total_demand = 0
            total_capacity = 0
            
            for diameter, result in self.optimization_results.items():
                if result:
                    total_bars += result['total_bins']
                    total_waste += result['total_waste']
                    total_demand += result['total_demand']
                    total_capacity += result['total_capacity']
            
            overall_waste_pct = (total_waste / total_capacity * 100) if total_capacity > 0 else 0
            
            # Data rows
            for diameter in sorted(self.optimization_results.keys()):
                result = self.optimization_results[diameter]
                if result:
                    data = [
                        diameter,
                        result['total_bins'],
                        round(result['total_waste'], 2),
                        round(result['waste_percentage'], 2),
                        round(result['total_demand'], 2)
                    ]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    row += 1
            
            # Total row
            total_data = ['TOTAL', total_bars, round(total_waste, 2), round(overall_waste_pct, 2), round(total_demand, 2)]
            for col, value in enumerate(total_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            row += 2
            
            # Detailed patterns for each diameter
            for diameter in sorted(self.optimization_results.keys()):
                result = self.optimization_results[diameter]
                if not result:
                    continue
                
                # Diameter header
                ws.merge_cells(f'A{row}:E{row}')
                cell = ws[f'A{row}']
                cell.value = f"DIAMETER: Ø{diameter}mm"
                cell.font = Font(bold=True, size=11)
                row += 1
                
                ws[f'A{row}'] = f"Demand: {result['total_demand']:.2f}m"
                row += 1
                ws[f'A{row}'] = f"Bars (Theoretical/Used): {result['theoretical_min']}/{result['total_bins']} bars"
                row += 1
                ws[f'A{row}'] = f"Waste: {result['total_waste']:.2f}m ({result['waste_percentage']:.2f}%)"
                row += 2
                
                # Pattern headers
                pattern_headers = ['Pattern', 'Bars', 'Cuts', 'Total(m)', 'Waste(m)']
                for col, header in enumerate(pattern_headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                row += 1
                
                # Get lengths for this diameter
                diameter_rebars = [r for r in self.rebar_list if r['diameter'] == diameter]
                lengths = [r['length'] for r in diameter_rebars]
                
                # Pattern data
                for idx, pattern_data in enumerate(result['used_patterns'], 1):
                    combo = pattern_data['combo']
                    count = pattern_data['count']
                    total = pattern_data['total']
                    waste = pattern_data['waste']
                    
                    # Build pattern description
                    cuts = []
                    for i, pieces in enumerate(combo):
                        if pieces > 0:
                            cuts.append(f"{pieces}×{lengths[i]:.2f}m")
                    pattern_desc = " + ".join(cuts)
                    
                    data = [f"Pattern {idx}", count, pattern_desc, round(total, 2), round(waste, 2)]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = border
                        if col == 3:  # Cuts column
                            cell.alignment = Alignment(horizontal='left')
                        else:
                            cell.alignment = Alignment(horizontal='center')
                    row += 1
                
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            
            # Save
            wb.save(filename)
            messagebox.showinfo("Success", f"Excel file saved:\n{filename}")
            self.status_label.config(text=f"● Excel saved: {os.path.basename(filename)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not save Excel file:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def save_as_pdf(self, filename):
        """Save report as PDF file"""
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror(
                "Error",
                "PDF export requires fpdf2!\n\n"
                "Install with: pip install fpdf2"
            )
            return
        
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Use monospace font for better alignment
            pdf.add_font('DejaVu', '', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', uni=True)
            pdf.set_font('DejaVu', '', 10)
            
            # Title
            pdf.set_font('DejaVu', '', 16)
            pdf.cell(0, 10, 'CUTTING PLAN - PRODUCTION INSTRUCTION', 0, 1, 'C')
            pdf.ln(5)
            
            # Info
            pdf.set_font('DejaVu', '', 10)
            pdf.cell(0, 6, f"Date: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 0, 1)
            pdf.cell(0, 6, f"Stock Bar Length: {self.stock_length}m", 0, 1)
            pdf.ln(5)
            
            # Calculate totals
            total_bars = 0
            total_waste = 0
            total_demand = 0
            total_capacity = 0
            
            for diameter, result in self.optimization_results.items():
                if result:
                    total_bars += result['total_bins']
                    total_waste += result['total_waste']
                    total_demand += result['total_demand']
                    total_capacity += result['total_capacity']
            
            overall_waste_pct = (total_waste / total_capacity * 100) if total_capacity > 0 else 0
            
            # Summary table
            pdf.set_font('DejaVu', '', 12)
            pdf.cell(0, 8, 'OVERALL SUMMARY - ALL DIAMETERS', 0, 1, 'C')
            pdf.ln(2)
            
            # Table header
            pdf.set_font('DejaVu', '', 9)
            pdf.set_fill_color(95, 149, 152)
            pdf.set_text_color(255, 255, 255)
            
            col_widths = [30, 25, 30, 25, 30]
            headers = ['DIAM(mm)', 'BARS', 'WASTE(m)', 'WASTE%', 'DEMAND(m)']
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, 1, 0, 'C', True)
            pdf.ln()
            
            # Table data
            pdf.set_text_color(0, 0, 0)
            pdf.set_fill_color(248, 249, 249)
            
            fill = False
            for diameter in sorted(self.optimization_results.keys()):
                result = self.optimization_results[diameter]
                if result:
                    pdf.cell(col_widths[0], 7, str(diameter), 1, 0, 'C', fill)
                    pdf.cell(col_widths[1], 7, str(result['total_bins']), 1, 0, 'C', fill)
                    pdf.cell(col_widths[2], 7, f"{result['total_waste']:.2f}", 1, 0, 'C', fill)
                    pdf.cell(col_widths[3], 7, f"{result['waste_percentage']:.2f}", 1, 0, 'C', fill)
                    pdf.cell(col_widths[4], 7, f"{result['total_demand']:.2f}", 1, 0, 'C', fill)
                    pdf.ln()
                    fill = not fill
            
            # Total row
            pdf.set_font('DejaVu', '', 9)
            pdf.cell(col_widths[0], 7, 'TOTAL', 1, 0, 'C', True)
            pdf.cell(col_widths[1], 7, str(total_bars), 1, 0, 'C', True)
            pdf.cell(col_widths[2], 7, f"{total_waste:.2f}", 1, 0, 'C', True)
            pdf.cell(col_widths[3], 7, f"{overall_waste_pct:.2f}", 1, 0, 'C', True)
            pdf.cell(col_widths[4], 7, f"{total_demand:.2f}", 1, 0, 'C', True)
            pdf.ln(10)
            
            # Detailed patterns for each diameter
            for diameter in sorted(self.optimization_results.keys()):
                result = self.optimization_results[diameter]
                if not result:
                    continue
                
                # Check if new page needed
                if pdf.get_y() > 250:
                    pdf.add_page()
                
                # Diameter header
                pdf.set_font('DejaVu', '', 12)
                pdf.cell(0, 8, f"DIAMETER: O{diameter}mm", 0, 1)
                pdf.set_font('DejaVu', '', 9)
                pdf.cell(0, 6, f"Demand: {result['total_demand']:.2f}m", 0, 1)
                pdf.cell(0, 6, f"Bars (Theoretical/Used): {result['theoretical_min']}/{result['total_bins']} bars", 0, 1)
                pdf.cell(0, 6, f"Waste: {result['total_waste']:.2f}m ({result['waste_percentage']:.2f}%)", 0, 1)
                pdf.ln(3)
                
                # Pattern table header
                pdf.set_fill_color(95, 149, 152)
                pdf.set_text_color(255, 255, 255)
                pattern_widths = [25, 20, 75, 25, 25]
                pattern_headers = ['Pattern', 'Bars', 'Cuts', 'Total(m)', 'Waste(m)']
                
                for i, header in enumerate(pattern_headers):
                    pdf.cell(pattern_widths[i], 7, header, 1, 0, 'C', True)
                pdf.ln()
                
                # Pattern data
                pdf.set_text_color(0, 0, 0)
                pdf.set_fill_color(248, 249, 249)
                
                # Get lengths for this diameter
                diameter_rebars = [r for r in self.rebar_list if r['diameter'] == diameter]
                lengths = [r['length'] for r in diameter_rebars]
                
                fill = False
                for idx, pattern_data in enumerate(result['used_patterns'], 1):
                    combo = pattern_data['combo']
                    count = pattern_data['count']
                    total = pattern_data['total']
                    waste = pattern_data['waste']
                    
                    # Build pattern description
                    cuts = []
                    for i, pieces in enumerate(combo):
                        if pieces > 0:
                            cuts.append(f"{pieces}x{lengths[i]:.2f}m")
                    pattern_desc = " + ".join(cuts)
                    
                    pdf.cell(pattern_widths[0], 6, f"Pattern {idx}", 1, 0, 'C', fill)
                    pdf.cell(pattern_widths[1], 6, str(count), 1, 0, 'C', fill)
                    pdf.cell(pattern_widths[2], 6, pattern_desc, 1, 0, 'L', fill)
                    pdf.cell(pattern_widths[3], 6, f"{total:.2f}", 1, 0, 'C', fill)
                    pdf.cell(pattern_widths[4], 6, f"{waste:.2f}", 1, 0, 'C', fill)
                    pdf.ln()
                    fill = not fill
                
                pdf.ln(5)
            
            # Warning note at the end
            pdf.set_font('DejaVu', '', 10)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 8, 'NOTE: Double-check all measurements before cutting.', 0, 1, 'C')
            pdf.ln(3)
            
            # Save PDF
            pdf.output(filename)
            messagebox.showinfo("Success", f"PDF file saved:\n{filename}")
            self.status_label.config(text=f"● PDF saved: {os.path.basename(filename)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not save PDF file:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def print_report(self):
        """Print report"""
        messagebox.showinfo(
            "Print",
            "To print the report, first save it using 'Save Report',\n"
            "then print the file from your text editor."
        )
    
    def copy_to_clipboard(self):
        """Copy report to clipboard"""
        content = self.results_text.get(1.0, tk.END)
        
        if not content.strip():
            messagebox.showwarning("Warning", "No report to copy!")
            return
        
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        messagebox.showinfo("Success", "Report copied to clipboard!")
        self.status_label.config(text="● Report copied to clipboard")


def main():
    """Main function to run the application"""
    root = tk.Tk()
    app = RebarOptimizerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()